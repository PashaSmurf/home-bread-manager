import string
import requests

from openpyxl import load_workbook

from telegram_bot.config.env_vars import ES_URL, EXCEL_DOWNLOAD_PATH


class Migration:
    @staticmethod
    def generate_table(cell_valle: str, sheet) -> list:
        alphabet = string.ascii_lowercase
        cell_list = cell_valle.split(':')
        start_row = alphabet.index(cell_list[0][0].lower())
        end_row = alphabet.index(cell_list[1][0].lower())
        start_col = int(cell_list[0][1:])
        end_col = int(cell_list[1][1:])

        res = []
        for row in range(start_col, end_col + 1):
            inner_list = []
            for col in range(start_row, end_row + 1):
                inner_list.append(sheet[row][col].value)
            res.append(inner_list)
        return res

    @staticmethod
    def excel_migration():
        workbook = load_workbook(filename=EXCEL_DOWNLOAD_PATH)
        for sheet_name in workbook.sheetnames:
            result = {}
            sheet = workbook[sheet_name]

            # title
            result['title'] = sheet['A2'].value

            # short description
            result['short_description'] = sheet['B2'].value

            # image_path
            result['image_path'] = '../images/' + sheet['C2'].value

            # metadata
            result['metadata'] = []
            result['metadata'].append(
                {'title': sheet['D1'].value, 'content': Migration.generate_table(sheet['D2'].value, sheet)})
            result['metadata'].append(
                {'title': sheet['E1'].value, 'content': Migration.generate_table(sheet['E2'].value, sheet)})
            result['metadata'].append(
                {'title': sheet['F1'].value, 'content': Migration.generate_table(sheet['F2'].value, sheet)})
            result['metadata'].append(
                {'title': sheet['G1'].value, 'content': Migration.generate_table(sheet['G2'].value, sheet)})
            result['metadata'].append(
                {'title': sheet['H1'].value, 'content': Migration.generate_table(sheet['H2'].value, sheet)})

            headers = {'Content-Type': 'application/json'}
            requests.post(ES_URL, headers=headers, json=result, verify=False)
